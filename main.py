from datetime import datetime
import random
from tkinter import messagebox
from tkinter.messagebox import askyesno
from customtkinter import *
import os
from PIL import Image, ImageTk, ImageFont

import os
from openpyxl import load_workbook
import pandas as pd
import sqlite3
import tkinter
from tkinter import ttk
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from barcode import EAN13
from barcode.writer import ImageWriter
import locale

locale.setlocale(locale.LC_TIME,'pt_BR.UTF-8')

dia_atual = datetime.now().strftime("%d/%m/%Y")

data_atual = datetime.now()
nome_mes = data_atual.strftime("%B/%Y")

# Obtém o diretório do usuário
diretorio = os.getcwd()

# Constrói o caminho completo do diretório dos itens
caminh_banco = os.path.join(diretorio, "Banco_de_Dados", "estoque.db")
caminh_plan_entrada = os.path.join(diretorio, "Banco_de_Dados", "Histórico Entrada.xlsx")
caminh_plan_saida = os.path.join(diretorio, "Banco_de_Dados", "Histórico Saída.xlsx")

caminh_layout = os.path.join(diretorio, "Banco_de_Dados", "LAYOUT TG.xlsx")

caminh_img_lat = os.path.join(diretorio, "assets", "side-img.png")
caminh_icon_user = os.path.join(diretorio, "assets", "user-icon.png")
caminh_icon_senha = os.path.join(diretorio, "assets", "password-icon.png")

# Definindo a Tela Principal
tela_login = CTk()
tela_login.geometry("{}x{}+0+0".format(tela_login.winfo_screenwidth(), tela_login.winfo_screenheight()))
tela_login.title("Estoque - Login")

#Criando variavel para coletar as dimensões de tamanho da tela
altura_tela = tela_login.winfo_screenheight()
largura_tela = tela_login.winfo_screenwidth()

# Carregando as imagens usando PIL
img_lat_pil = Image.open(caminh_img_lat)
icone_user_pil = Image.open(caminh_icon_user)
icone_senha_pil = Image.open(caminh_icon_senha)

#Puxando as imagens para usar na interface
img_lat = CTkImage(dark_image=img_lat_pil, light_image=img_lat_pil, size=(int(0.4*largura_tela), altura_tela))
icone_user = CTkImage(dark_image=icone_user_pil, light_image=icone_user_pil, size=(20, 20))
icone_senha = CTkImage(dark_image=icone_senha_pil, light_image=icone_senha_pil, size=(17, 17))
font_path = "Arial.ttf"

def sair_do_app():
    ans = askyesno(title='Sair', message='Tem certeza que quer Sair?')
    if ans:
        sys.exit()

def preencher_descricao(event):
    df_layout = pd.read_excel(caminh_layout)
    codigo = int(code_entry.get()) # Pega o código digitado
    descricao = df_layout[df_layout['CODE'] == codigo]['DESCRIÇÃO'].values # Busca a descrição correspondente
    if len(descricao) > 0:
        nome_entry.delete(0, tkinter.END)
        nome_entry.insert(0, descricao[0])
    else:
        nome_entry.delete(0, tkinter.END)

def autenticacao():
    global login, senha, tela_menu
    login = logn.get()
    senha = passw.get()

    if (login == "admin" and senha == "admin"):

        tela_login.withdraw()

        tela_menu = CTkToplevel()
        tela_menu.geometry("{}x{}+0+0".format(largura_tela,altura_tela))
        tela_menu.title('Estoque - Menu')
        tela_menu.protocol("WM_DELETE_WINDOW", sair_do_app)

        CTkLabel(master=tela_menu, text="", image=img_lat).pack(expand=True, side="left")

        frame_cabecalho_menu = CTkFrame(master=tela_menu, width=int(0.6 * largura_tela), height=int(0.14 * altura_tela))
        frame_cabecalho_menu.pack(expand=True, side="top")

        CTkLabel(master=frame_cabecalho_menu, text="------------------------------------- Estoque -------------------------------------", text_color="#00009C", anchor="w", justify="center", font=("Arial Bold", 24)).pack(anchor="w", pady=(0,0), padx=(0, 0))

        scrollable_frame_menu = CTkScrollableFrame(master=tela_menu, width=int(0.6 * largura_tela), height=int(0.85 * altura_tela), fg_color="#ffffff")
        scrollable_frame_menu.pack(expand=True, side="right")

        frame_menu = CTkFrame(master=scrollable_frame_menu, fg_color="#ffffff")
        frame_menu.pack(expand=True, anchor="w", pady=(10, 0))

        CTkLabel(master=frame_menu, text="  Verificar Estoque:", text_color="#00009C", anchor="w", justify="center", font=("Arial Bold", 14), compound="left").pack(anchor="w", pady=(35, 0), padx=(int(0.2*frame_menu.winfo_screenwidth()), 0))
        CTkButton(master=frame_menu, text="Estoque", fg_color="#00009C", hover_color="#E44982", font=("Arial Bold", 12), text_color="#ffffff", width=225, command=ver_estoque).pack(anchor="w", pady=(15, 0), padx=(int(0.2*frame_menu.winfo_screenwidth()), 0))

        CTkLabel(master=frame_menu, text="  Ver Gráficos:", text_color="#00009C", anchor="w", justify="center", font=("Arial Bold", 14), compound="left").pack(anchor="w", pady=(35, 0), padx=(int(0.2*frame_menu.winfo_screenwidth()), 0))
        CTkButton(master=frame_menu, text="Gráficos", fg_color="#00009C", hover_color="#E44982", font=("Arial Bold", 12), text_color="#ffffff", width=225, command=menu_graficos).pack(anchor="w", pady=(15, 0), padx=(int(0.2*frame_menu.winfo_screenwidth()), 0))

        CTkLabel(master=frame_menu, text="  Acompanhamento de Entrada/Saída:", text_color="#00009C", anchor="w", justify="center", font=("Arial Bold", 14), compound="left").pack(anchor="w", pady=(35, 0), padx=(int(0.2*frame_menu.winfo_screenwidth()), 0))
        CTkButton(master=frame_menu, text="Ver Acompanhamento", fg_color="#00009C", hover_color="#E44982", font=("Arial Bold", 12), text_color="#ffffff", width=225, command=menu_acomp).pack(anchor="w", pady=(15, 0), padx=(int(0.2*frame_menu.winfo_screenwidth()), 0))

        CTkLabel(master=frame_menu, text="  Encontrar Item:", text_color="#00009C", anchor="w", justify="center", font=("Arial Bold", 14), compound="left").pack(anchor="w", pady=(35, 0), padx=(int(0.2*frame_menu.winfo_screenwidth()), 0))
        CTkButton(master=frame_menu, text="Layout", fg_color="#00009C", hover_color="#E44982", font=("Arial Bold", 12), text_color="#ffffff", width=225, command=encontrar_item).pack(anchor="w", pady=(15, 0), padx=(int(0.2*frame_menu.winfo_screenwidth()), 0))

    else:
        messagebox.showerror("Erro!", "Login ou Senha incorretos")


def ver_estoque():

    global tela_estoque, treeviewF, code_entry, nome_entry, qtd_entry, un_entry, preco_entry, code_entry_pes

    tela_menu.withdraw()
    
    tela_estoque = CTkToplevel()
    tela_estoque.geometry("{}x{}+0+0".format(largura_tela, altura_tela))
    tela_estoque.title("Estoque - Estoque Físico")
    tela_estoque.protocol("WM_DELETE_WINDOW", sair_do_app)

    frame_estoque = tkinter.Frame(tela_estoque, width= int(0.7 * largura_tela), height=int(altura_tela))
    frame_estoque.pack_propagate(0)
    frame_estoque.pack(expand=True, side="right")

    frame_treeview_estoque = tkinter.Frame(frame_estoque, width= int(0.7 * largura_tela), height=int(altura_tela))
    frame_treeview_estoque.pack_propagate(0)
    frame_treeview_estoque.pack(expand=True, side="right")

    # Criar a Treeview
    treeviewF = ttk.Treeview(frame_treeview_estoque, columns=("Nome", "Descrição"), show="headings")
    treeviewF.heading("Nome", text="Nome")
    treeviewF.heading("Descrição", text="Descrição")

    # Adicionar Scrollbars
    scrollbar_y = ttk.Scrollbar(frame_treeview_estoque, orient="vertical", command=treeviewF.yview)
    scrollbar_y.pack(side="right", fill="y")
    treeviewF.configure(yscrollcommand=scrollbar_y.set)

    scrollbar_x = ttk.Scrollbar(frame_treeview_estoque, orient="horizontal", command=treeviewF.xview)
    scrollbar_x.pack(side="bottom", fill="x")
    treeviewF.configure(xscrollcommand=scrollbar_x.set)

    frame_widgets_estoque = CTkScrollableFrame(tela_estoque, width=int(0.3 * largura_tela), height=altura_tela)
    frame_widgets_estoque.pack(side='left', fill="both", expand=True)

    # Widgets no FrameFB
    ttk.Label(frame_widgets_estoque)
    code_entry = ttk.Entry(frame_widgets_estoque)
    code_entry.insert(0, "Código")
    code_entry.bind("<FocusIn>", lambda e: code_entry.delete('0','end'))
    code_entry.bind('<KeyRelease>', preencher_descricao)
    code_entry.pack(pady=50, padx=20, fill="x")

    nome_entry = ttk.Entry(frame_widgets_estoque)
    nome_entry.insert(0, "Descrição")
    nome_entry.bind("<FocusIn>", lambda e: nome_entry.delete('0','end'))
    nome_entry.pack(pady=10, padx=20, fill="x")

    qtd_entry = ttk.Entry(frame_widgets_estoque)
    qtd_entry.insert(0,"Quantidade")
    qtd_entry.bind("<FocusIn>", lambda e: qtd_entry.delete('0','end'))
    qtd_entry.pack(pady=10, padx=20, fill="x")

    un_entry = ttk.Entry(frame_widgets_estoque)
    un_entry.insert(0,"Unidade")
    un_entry.bind("<FocusIn>", lambda e: un_entry.delete('0','end'))
    un_entry.pack(pady=10, padx=20, fill="x")

    preco_entry = ttk.Entry(frame_widgets_estoque)
    preco_entry.insert(0,"Preço")
    preco_entry.bind("<FocusIn>", lambda e: preco_entry.delete('0','end'))
    preco_entry.pack(pady=10, padx=20, fill="x")

    botao = ttk.Button(frame_widgets_estoque, text="Inserir", command=inserir_item)
    botao.pack(pady=10, padx=20, fill="x")

    separator = ttk.Separator(frame_widgets_estoque)
    separator.pack(pady=10, padx=20, fill="x")

    code_entry_pes = ttk.Entry(frame_widgets_estoque)
    code_entry_pes.insert(0, "Código")
    code_entry_pes.bind("<FocusIn>", lambda e: code_entry_pes.delete('0','end'))
    code_entry_pes.pack(pady=10, padx=20, fill="x")

    botao_pes = ttk.Button(frame_widgets_estoque, text="Pesquisar",command=pesquisar_produto)
    botao_pes.pack(pady=10, padx=20, fill="x")
        
    botao_vw = ttk.Button(frame_widgets_estoque, text="Limpar Pesquisa",command=limpar_pesquisa)
    botao_vw.pack(pady=10, padx=20, fill="x")

    botao_vw_cdb = ttk.Button(frame_widgets_estoque, text="Ver Código de Barras", command=exibir_codigo_barras_selecionado)
    botao_vw_cdb.pack(pady=10, padx=20, fill="x")

    botao_pesq_cdb= ttk.Button(frame_widgets_estoque, text="Pesquisar Código de Barras", command=pesquisar_codigo_de_barras)
    botao_pesq_cdb.pack(pady=10, padx=20, fill="x")

    botao_back = ttk.Button(frame_widgets_estoque, text="Voltar", command=voltar_menu)
    botao_back.pack(pady=10, padx=20, fill="x")
    try:
            conn = sqlite3.connect(caminh_banco)
            cursor = conn.cursor()

            cursor.execute("SELECT code, descricao, qtd, un, preco, codigo_barras FROM fisico")
            rows = cursor.fetchall()

            treeviewF.delete(*treeviewF.get_children())

            cols = ["CODE", "DESCRIÇÃO", "QUANTIDADE", "UNIDADE","PREÇO","CÓDIGO DE BARRAS"]
            treeviewF["columns"] = cols

            for col_name in cols:
                treeviewF.heading(col_name, text=col_name)
                treeviewF.column(col_name, anchor=tkinter.CENTER)

            #Adicionar as tags de estilo para controle de Estoque
            treeviewF.tag_configure('light_green', background='#90EE90')  # Verde claro
            treeviewF.tag_configure('light_yellow', background='yellow')  # Amarelo (mesma cor)
            treeviewF.tag_configure('light_red', background='#FF7F7F')    # Vermelho claro
            for row in rows:
                quantidade = row[2]  #Coluna QUANTIDADE
                if quantidade >100:
                    treeviewF.insert('', tkinter.END, values=row, tags=('light_green',))
                elif quantidade <= 100 and quantidade > 60:
                    treeviewF.insert('', tkinter.END, values=row, tags=('light_green',))
                elif quantidade <= 60 and quantidade > 30:
                    treeviewF.insert('', tkinter.END, values=row, tags=('light_yellow',))
                elif quantidade <= 30:
                    treeviewF.insert('', tkinter.END, values=row, tags=('light_red',))
                else:
                    treeviewF.insert('', tkinter.END, values=row)
            

            treeviewF.pack(expand=True, fill="both")  
            messagebox.showinfo("Aviso!","Dados carregados com sucesso!")
    except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar dados do banco de dados: {str(e)}")
    finally:
            if conn:
                conn.close()

def inserir_item():

        # Recuperar os valores dos campos de entrada
        cdg = code_entry.get()
        prod = nome_entry.get()
        qtd = qtd_entry.get()
        un = un_entry.get()
        preco = preco_entry.get()

        # Verificar se os campos estão preenchidos corretamente
        if not all([cdg, prod, qtd, un, preco]) or prod == "Descrição" or qtd == "Quantidade" or cdg == "Código" or un == "Unidade" or preco == "Preço":
            messagebox.showwarning("Aviso!", "Preencha todos os campos")
        else:
            # Converter os valores para os tipos corretos
            qtd = int(qtd)
            preco = float(preco)

            # Conectar ao banco de dados
            conn = sqlite3.connect(caminh_banco)
            cursor = conn.cursor()

            # Verificar se o item já existe no banco de dados

            cursor.execute("SELECT qtd, preco FROM fisico WHERE code = ?", (cdg,))
            resultado = cursor.fetchone()

            if resultado:
                # Se o item já existe, atualizar a quantidade e o preço
                qtd_atual = resultado[0]
                preco_atual = resultado[1]

                nova_qtd = qtd_atual + int(qtd)  # Certifique-se de que qtd seja um número
                novo_preco = preco_atual + float(preco)  # Certifique-se de que preco seja um número

                # Atualiza tanto a quantidade quanto o preço
                cursor.execute("UPDATE fisico SET qtd = ?, preco = ? WHERE code = ?", (nova_qtd, novo_preco, cdg))
                conn.commit()
                messagebox.showinfo("Aviso!", "Quantidade e preço atualizados com sucesso!")

            else:
                # Se o item não existe, gerar o código de barras e inseri-lo

                df_layout = pd.read_excel(caminh_layout, sheet_name="LAYOUT")
                df_layout['CODE'] = df_layout['CODE'].astype(str)
                df_layout = df_layout[df_layout['CODE'] == cdg]

                codigo = df_layout['CDBAR'].iloc[0]

                codigo_barras = gerar_codigo_barras(str(codigo))
                
                dados = (str(cdg), prod, qtd, un, preco, codigo_barras)
                cursor.execute("INSERT INTO fisico (code, descricao, qtd, un, preco, codigo_barras) VALUES (?, ?, ?, ?, ?, ?)", dados)
                conn.commit()
                
                # Inserir os dados na tabela de entrada
                dados_entrada = (str(cdg), prod, qtd, un, preco, codigo_barras, dia_atual, nome_mes)
                cursor.execute("INSERT INTO entrada (code, descricao, qtd, un, preco, codigo_barras, data_entrada, mes_entrada) VALUES (?, ?, ?, ?, ?, ?, ?, ?)", dados_entrada)
                conn.commit()

                # Inserir os dados na planilha do Excel
                workbook = load_workbook(caminh_plan_entrada)
                sheet = workbook['Hist']
                sheet.append(dados_entrada)
                workbook.save(caminh_plan_entrada)

                messagebox.showinfo("Aviso!", "Item inserido com sucesso!")
            
            # Fechar a conexão com o banco de dados
            conn.close()

            # Limpar os widgets de entrada após a inserção
            code_entry.delete(0, tkinter.END)
            nome_entry.delete(0, tkinter.END)
            qtd_entry.delete(0, tkinter.END)
            un_entry.delete(0, tkinter.END)
            preco_entry.delete(0, tkinter.END)

            # Reinserir os valores padrão
            code_entry.insert(0, "Código")
            nome_entry.insert(0, "Nome do Produto")
            qtd_entry.insert(0, "Quantidade")
            un_entry.insert(0, "Unidade")
            preco_entry.insert(0, "Preço")



def gerar_sequencia_aleatoria():
    sequencia = ''.join(str(random.randint(0, 9)) for _ in range(12))
    return str(sequencia)

def codigo_barras_existe_no_bd(codigo_barras):
    conn = sqlite3.connect(caminh_banco)   
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) FROM fisico WHERE codigo_barras = ?", (codigo_barras,))
    quantidade = cursor.fetchone()[0]
    conn.close()
    return quantidade > 0

def gerar_codigo_barras_unico():
    while True:
        nova_sequencia = gerar_sequencia_aleatoria()
        if not codigo_barras_existe_no_bd(nova_sequencia):
            return str(nova_sequencia)

def gerar_codigo_barras(codigo):
    return str(EAN13(codigo, writer=ImageWriter()))


def gerar_codigo_barras_img(codigo):
    # Defina o caminho da fonte no seu projeto
    font_path = os.path.join(diretorio, 'arial.ttf')
    
    # Configure o ImageWriter com o font_type
    writer = ImageWriter()
    writer.font_path = font_path  # Defina o caminho da fonte no ImageWriter
    
    # Gere o código de barras
    return EAN13(codigo, writer=writer)


def pesquisar_produto():
    cdg_pesq = code_entry_pes.get().lower()

    try:
        conn = sqlite3.connect(caminh_banco)
        cursor = conn.cursor()

        cursor.execute(f"SELECT code, descricao, qtd, un, preco, codigo_barras FROM fisico WHERE code = '{cdg_pesq}'")
        rows = cursor.fetchall()

        treeviewF.delete(*treeviewF.get_children())

        cols = ["CODE", "DESCRIÇÃO", "QUANTIDADE", "UNIDADE","PREÇO","CÓDIGO DE BARRAS"]
        treeviewF["columns"] = cols

        for col_name in cols:
            treeviewF.heading(col_name, text=col_name)
            treeviewF.column(col_name, anchor=tkinter.CENTER)

        #Adicionar as tags de estilo para controle de Estoque
        treeviewF.tag_configure('light_green', background='#90EE90')  # Verde claro
        treeviewF.tag_configure('light_yellow', background='yellow')  # Amarelo (mesma cor)
        treeviewF.tag_configure('light_red', background='#FF7F7F')    # Vermelho claro
        for row in rows:
            quantidade = row[2]  #Coluna QUANTIDADE
            if quantidade >100:
                treeviewF.insert('', tkinter.END, values=row, tags=('light_green',))
            elif quantidade <= 100 and quantidade > 60:
                treeviewF.insert('', tkinter.END, values=row, tags=('light_green',))
            elif quantidade <= 60 and quantidade > 30:
                treeviewF.insert('', tkinter.END, values=row, tags=('light_yellow',))
            elif quantidade <= 30:
                treeviewF.insert('', tkinter.END, values=row, tags=('light_red',))
            else:
                treeviewF.insert('', tkinter.END, values=row)

    except Exception as e:
        messagebox.showerror("Erro", f"Item não encontrado: {str(e)}")
    finally:
        if conn:
            conn.close()

def limpar_pesquisa():
    try:
            conn = sqlite3.connect(caminh_banco)
            cursor = conn.cursor()

            cursor.execute("SELECT code, descricao, qtd, un, preco, codigo_barras FROM fisico")
            rows = cursor.fetchall()

            treeviewF.delete(*treeviewF.get_children())

            cols = ["CODE", "DESCRIÇÃO", "QUANTIDADE", "UNIDADE","PREÇO","CÓDIGO DE BARRAS"]
            treeviewF["columns"] = cols

            for col_name in cols:
                treeviewF.heading(col_name, text=col_name)
                treeviewF.column(col_name, anchor=tkinter.CENTER)

            #Adicionar as tags de estilo para controle de Estoque
            treeviewF.tag_configure('light_green', background='#90EE90')  # Verde claro
            treeviewF.tag_configure('light_yellow', background='yellow')  # Amarelo (mesma cor)
            treeviewF.tag_configure('light_red', background='#FF7F7F')    # Vermelho claro
            for row in rows:
                quantidade = row[2]  #Coluna QUANTIDADE
                if quantidade >100:
                    treeviewF.insert('', tkinter.END, values=row, tags=('light_green',))
                elif quantidade <= 100 and quantidade > 60:
                    treeviewF.insert('', tkinter.END, values=row, tags=('light_green',))
                elif quantidade <= 60 and quantidade > 30:
                    treeviewF.insert('', tkinter.END, values=row, tags=('light_yellow',))
                elif quantidade <= 30:
                    treeviewF.insert('', tkinter.END, values=row, tags=('light_red',))
                else:
                    treeviewF.insert('', tkinter.END, values=row)

            treeviewF.pack(expand=True, fill="both")  
            messagebox.showinfo("Aviso!","Dados carregados com sucesso!")
    except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar dados do banco de dados: {str(e)}")
    finally:
            if conn:
                conn.close()

def exibir_codigo_barras_selecionado():
    #Obtém o item selecionado na treeview
    item_selecionado = treeviewF.selection()

    if item_selecionado:
        #Obtém o código de barras associado ao item selecionado
        codigo_barras = treeviewF.item(item_selecionado, 'values')[5]
        produto = treeviewF.item(item_selecionado, 'values')[0]

        # Exibe o código de barras na nova janela
        codigo = gerar_codigo_barras_img(codigo_barras)
        codigo_imagem = ImageTk.PhotoImage(codigo.render())

        tela_estoque.withdraw()

        global tela_cod_de_barras

        tela_cod_de_barras = CTkToplevel()
        tela_cod_de_barras.geometry("{}x{}+0+0".format(largura_tela, altura_tela))
        tela_cod_de_barras.title("Estoque - Código de Barras")
        tela_cod_de_barras.protocol("WM_DELETE_WINDOW", sair_do_app)

        CTkLabel(master=tela_cod_de_barras, text="", image=img_lat).pack(expand=True, side="left")

        #Traz a imagem gerada do Código de Barras
        frame_cod_de_barras = CTkScrollableFrame(master=tela_cod_de_barras, width= int(0.6 * largura_tela), height=int(altura_tela), fg_color="#ffffff")
        frame_cod_de_barras.pack_propagate(0)
        frame_cod_de_barras.pack(expand=True, side="right")

        frame_info_cod_de_barras = ttk.Frame(frame_cod_de_barras)
        frame_info_cod_de_barras.grid(row=0, column=0, padx=10, pady=10)

        CTkLabel(master=frame_info_cod_de_barras, text="-----------------", text_color="#ffffff", anchor="w", justify="left", font=("Arial Bold", 18)).grid(row=1, column=0, padx=10, pady=10)
        CTkLabel(master=frame_info_cod_de_barras, text=f"Estoque - Código de Barras - {produto} ", text_color="#00009C", anchor="w", justify="left", font=("Arial Bold", 18)).grid(row=1, column=1, padx=10, pady=10)   
        CTkLabel(master=frame_info_cod_de_barras, text="--------------", text_color="#ffffff", anchor="w", justify="left", font=("Arial Bold", 18)).grid(row=1, column=2, padx=10, pady=10)        
        CTkButton(master=frame_info_cod_de_barras, text="Voltar", fg_color="#00009C", hover_color="#E44982", font=("Arial Bold", 12), text_color="#ffffff", width=50, command=voltar_tela_cdb).grid(row=1, column=3, padx=10, pady=10)

        frame_img_cd_barras = ttk.Label(frame_info_cod_de_barras, image=codigo_imagem)
        frame_img_cd_barras.image = codigo_imagem
        frame_img_cd_barras.grid(row=2, column=1, padx=10, pady=10)
        entry_code = CTkEntry(master=frame_info_cod_de_barras, text_color="#00009C", justify="center", width=300, font=("Arial Bold", 18))
        entry_code.grid(row=3, column=1, padx=10, pady=10)
        entry_code.insert(0, codigo_barras)
        entry_code.configure(state="readonly")
        CTkButton(master=frame_info_cod_de_barras, text="Imprimir", fg_color="#00009C", hover_color="#E44982", font=("Arial Bold", 12), text_color="#ffffff", width=50, command=lambda: salvar_imagem(produto, codigo_imagem, codigo_barras)).grid(row=4, column=1, padx=10, pady=10)

    else:
        messagebox.showinfo("Aviso!", "escolha o item que você gostaria de acessar o Código de barras")

def salvar_imagem(produto, imagem, cbar):
    diretorio_projeto = os.getcwd()
    print(diretorio_projeto)
    file_path = f"{diretorio_projeto}\códigos_de_barras\{produto} - {cbar}.png"

    #Converte a imagem
    imagem_pil = ImageTk.getimage(imagem)

    #Salva a imagem no diretório
    imagem_pil.save(file_path)

    messagebox.showinfo("Sucesso", f"Imagem do código de barras salva em {file_path}")

def voltar_tela_cdb():
    tela_cod_de_barras.withdraw()
    tela_estoque.deiconify()

def pesquisar_codigo_de_barras():
    global tela_saida, frame_saida
    global cd_entry
    tela_estoque.withdraw()

    tela_saida = CTkToplevel()
    tela_saida.geometry("{}x{}+0+0".format(largura_tela, altura_tela))
    tela_saida.title("Estoque")
    tela_saida.protocol("WM_DELETE_WINDOW", sair_do_app)

    CTkLabel(master=tela_saida, text="", image=img_lat).pack(expand=True, side="left")

    frame_cabecalho_saida = CTkFrame(master=tela_saida, width=int(0.6 * largura_tela), height=int(0.14 * altura_tela))
    frame_cabecalho_saida.pack(expand=True, side="top")

    CTkButton(master=frame_cabecalho_saida, text="Voltar", fg_color="#00009C", hover_color="#E44982", font=("Arial Bold", 12), text_color="#ffffff", width=50, command=voltar_pesq_cdb).pack(anchor="w", pady=(0,0), padx=(int(0.55*frame_cabecalho_saida.winfo_screenwidth()), 0))
    CTkLabel(master=frame_cabecalho_saida, text="Estoque - Pesquisar Código de Barras", text_color="#00009C", anchor="w", justify="left", font=("Arial Bold", 18)).pack(anchor="w", pady=(10,0), padx=(0, 0))

    frame_saida_scroll = CTkScrollableFrame(master=tela_saida, width= int(0.6 * largura_tela), height=int(0.85 * altura_tela), fg_color="#ffffff")
    frame_saida_scroll.pack(expand=True, side="right")

    frame_saida = ttk.Frame(frame_saida_scroll)
    frame_saida.grid(row=0, column=0, padx=10, pady=20)

    CTkLabel(master=frame_saida, text="-------------------", text_color="#ffffff", anchor="w", justify="left", font=("Arial Bold", 18)).grid(row=1, column=0, padx=10, pady=20)
    CTkLabel(master=frame_saida, text="Estoque - Pesquisar Código de Barras", text_color="#ffffff", anchor="w", justify="left", font=("Arial Bold", 18)).grid(row=1, column=1, padx=10, pady=20)
    cd_entry = CTkEntry(master=frame_saida, width=225, fg_color="#EEEEEE", border_color="#00009C", border_width=1, text_color="#000000")
    cd_entry.grid(row=2, column=1, padx=10, pady=20)
    CTkButton(master=frame_saida, text="Escanear", fg_color="#00009C", hover_color="#E44982", font=("Arial Bold", 12), text_color="#ffffff", width=50, command=escanear_codigo_de_barras).grid(row=3, column=1, padx=10, pady=20)

def escanear_codigo_de_barras():
    global codigo_pesquisado, produto_cde
    # Capturar a entrada
    codigo_pesquisado = cd_entry.get()

    try:
            diretorio_projeto = os.getcwd()
            conn = sqlite3.connect(caminh_banco)
            cursor = conn.cursor()

            cursor.execute(f"SELECT code FROM fisico WHERE codigo_barras = '{codigo_pesquisado}'")
            produto_cd = cursor.fetchone()
            conn.close()
            if produto_cd:
                produto_cde = produto_cd[0]

                imagem_cd_pesquisado = f"{diretorio_projeto}/códigos_de_barras/{produto_cde} - {codigo_pesquisado}.png"  
                imagem_cd_pesquisado_pillow = Image.open(imagem_cd_pesquisado)

                imagem_cd_pesquisado_tk = ImageTk.PhotoImage(imagem_cd_pesquisado_pillow)

                frame_saida_cdbarras = ttk.Label(frame_saida, image=imagem_cd_pesquisado_tk)
                frame_saida_cdbarras.image = imagem_cd_pesquisado_tk  
                frame_saida_cdbarras.grid(row=3, column=1, padx=10, pady=10)
                CTkLabel(master=frame_saida, text=f"Produto - {produto_cde}", text_color="#00009C", anchor="w", justify="left", font=("Arial Bold", 18)).grid(row=4, column=1, padx=10, pady=10)
                
                CTkButton(master=frame_saida, text="Ver Informações", fg_color="#00009C", font=("Arial Bold", 12), text_color="#ffffff", width=50, command=mostrar_info_produto_pesquisado).grid(row=5, column=1, pady=10)
                CTkButton(master=frame_saida, text="Excluir", fg_color="#00009C", font=("Arial Bold", 12), text_color="#ffffff", width=50, command=inserir_info_para_saida).grid(row=6, column=1, padx=10, pady=10)          
            else:
                messagebox.showerror("Aviso!", "Item não encontrado")

    except Exception as e:
        messagebox.showerror("Erro!", f"{str(e)}")

def mostrar_info_produto_pesquisado():
    tela_saida.withdraw()
    tela_estoque.deiconify()
    try:
        conn = sqlite3.connect(caminh_banco)
        cursor = conn.cursor()

        cursor.execute(f"SELECT code, descricao, qtd, un, preco, codigo_barras FROM fisico WHERE code = '{produto_cde}'")
        rows = cursor.fetchall()

        treeviewF.delete(*treeviewF.get_children())

        cols = ["CODE", "DESCRIÇÃO", "QUANTIDADE", "UNIDADE","PREÇO","CÓDIGO DE BARRAS"]
        treeviewF["columns"] = cols

        for col_name in cols:
            treeviewF.heading(col_name, text=col_name)
            treeviewF.column(col_name, anchor=tkinter.CENTER)

        #Adicionar as tags de estilo para controle de Estoque
        treeviewF.tag_configure('light_green', background='#90EE90')  # Verde claro
        treeviewF.tag_configure('light_yellow', background='yellow')  # Amarelo (mesma cor)
        treeviewF.tag_configure('light_red', background='#FF7F7F')    # Vermelho claro
        for row in rows:
            quantidade = row[2]  #Coluna QUANTIDADE
            if quantidade >100:
                    treeviewF.insert('', tkinter.END, values=row, tags=('light_green',))
            elif quantidade <= 100 and quantidade > 60:
                    treeviewF.insert('', tkinter.END, values=row, tags=('light_green',))
            elif quantidade <= 60 and quantidade > 30:
                treeviewF.insert('', tkinter.END, values=row, tags=('light_yellow',))
            elif quantidade <= 30:
                treeviewF.insert('', tkinter.END, values=row, tags=('light_red',))
            else:
                treeviewF.insert('', tkinter.END, values=row)


    except Exception as e:
        messagebox.showerror("Erro", f"Item não encontrado: {str(e)}")
    finally:
        if conn:
            conn.close()

def voltar_pesq_cdb():
    tela_saida.withdraw()
    tela_estoque.deiconify()
    try:
            conn = sqlite3.connect(caminh_banco)
            cursor = conn.cursor()

            cursor.execute("SELECT code, descricao, qtd, un, preco, codigo_barras FROM fisico")
            rows = cursor.fetchall()

            treeviewF.delete(*treeviewF.get_children())

            cols = ["CODE", "DESCRIÇÃO", "QUANTIDADE", "UNIDADE","PREÇO","CÓDIGO DE BARRAS"]
            treeviewF["columns"] = cols

            for col_name in cols:
                treeviewF.heading(col_name, text=col_name)
                treeviewF.column(col_name, anchor=tkinter.CENTER)

            #Adicionar as tags de estilo para controle de Estoque
            treeviewF.tag_configure('light_green', background='#90EE90')  # Verde claro
            treeviewF.tag_configure('light_yellow', background='yellow')  # Amarelo (mesma cor)
            treeviewF.tag_configure('light_red', background='#FF7F7F')    # Vermelho claro
            for row in rows:
                quantidade = row[2]  #Coluna QUANTIDADE
                if quantidade >100:
                    treeviewF.insert('', tkinter.END, values=row, tags=('light_green',))
                elif quantidade <= 100 and quantidade > 60:
                    treeviewF.insert('', tkinter.END, values=row, tags=('light_green',))
                elif quantidade <= 60 and quantidade > 30:
                    treeviewF.insert('', tkinter.END, values=row, tags=('light_yellow',))
                elif quantidade <= 30:
                    treeviewF.insert('', tkinter.END, values=row, tags=('light_red',))
                else:
                    treeviewF.insert('', tkinter.END, values=row)

            treeviewF.pack(expand=True, fill="both")
            messagebox.showinfo("Aviso!","Dados carregados com sucesso!")
    except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar dados do banco de dados: {str(e)}")
    finally:
            if conn:
                conn.close()

def inserir_info_para_saida():
    global quant,area
    CTkLabel(master=frame_saida, text="Digite a Quantidade para saída", text_color="#00009C", anchor="w", justify="left", font=("Arial Bold", 14)).grid(row=7, column=1, padx=10, pady=10)        
    quant = CTkEntry(master=frame_saida, width=225, fg_color="#EEEEEE", border_color="#00009C", border_width=1, text_color="#000000")
    quant.grid(row=8, column=1, padx=10, pady=10)
    CTkLabel(master=frame_saida, text="Digite a Área", text_color="#00009C", anchor="w", justify="left", font=("Arial Bold", 14)).grid(row=9, column=1, padx=10, pady=10)        
    area = CTkEntry(master=frame_saida, width=225, fg_color="#EEEEEE", border_color="#00009C", border_width=1, text_color="#000000")
    area.grid(row=10, column=1, padx=10, pady=10)
    CTkButton(master=frame_saida, text="Confirmar", fg_color="#00009C", font=("Arial Bold", 12), text_color="#ffffff", width=50, command=retirar_item).grid(row=11, column=1, padx=10, pady=10)        

def retirar_item():
    codigo = codigo_pesquisado
    ar = area.get()

    if not quant.get():
        messagebox.showwarning("Aviso", "Por favor, insira a quantidade desejada para excluir.")
        return
    try:
        quantidade = int(quant.get())
    except ValueError:
        messagebox.showerror("Erro", "A quantidade inserida não é um número válido.")
        return
    
    try:
        conn = sqlite3.connect(caminh_banco)
        cursor = conn.cursor()

        # Obter a quantidade atual do banco de dados
        cursor.execute(f"SELECT code, descricao, qtd, un, preco, codigo_barras FROM fisico WHERE codigo_barras = '{codigo}'")
        resultado_qtd = cursor.fetchone()

        if resultado_qtd:
            code = str(resultado_qtd[0])
            produto = str(resultado_qtd[1])
            qtd =  int(resultado_qtd[2])
            un =  str(resultado_qtd[3])
            preco = str(resultado_qtd[4])
            cd_bar =  str(resultado_qtd[5])

            if quantidade < qtd:
                # Atualizar a quantidade no banco de dados
                nova_quantidade = qtd - quantidade
                cursor.execute(f"UPDATE fisico SET qtd = {nova_quantidade} WHERE codigo_barras = '{codigo}'")
                conn.commit()

                dados_saida = (str(code),produto,int(quantidade),un,float(preco), ar,cd_bar, dia_atual, nome_mes)
                

                conn = sqlite3.connect(caminh_banco)
                cursor = conn.cursor()
                cursor.execute("INSERT INTO saida (code, descricao, qtd, un, preco, area, codigo_barras, data_saida,mes_saida) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)", dados_saida)
                conn.commit()
                conn.close()

                workbook = load_workbook(caminh_plan_saida)

                # Selecionar a planilha ativa (ou uma específica pelo nome)
                sheet = workbook['Hist']

                # Inserir os dados na próxima linha disponível usando append
                sheet.append(dados_saida)

                # Salvar as alterações na planilha
                workbook.save(caminh_plan_saida)


                messagebox.showinfo("Sucesso", f"{quantidade} removidos com sucesso.")
            elif quantidade == qtd:
                # Deletar completamente o item do banco de dados
                cursor.execute(f"DELETE FROM fisico WHERE codigo_barras = '{codigo}'")
                conn.commit()

                dados_saida = (str(code),produto,int(quantidade),un,float(preco), ar,cd_bar, dia_atual, nome_mes)

                conn = sqlite3.connect(caminh_banco)
                cursor = conn.cursor()
                cursor.execute("INSERT INTO saida (code, descricao, qtd, un, preco, area,codigo_barras, data_saida, mes_saida) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)", dados_saida)
                conn.commit()
                conn.close()

                workbook = load_workbook(caminh_plan_saida)

                # Selecionar a planilha ativa (ou uma específica pelo nome)
                sheet = workbook['Hist']

                # Inserir os dados na próxima linha disponível usando append
                sheet.append(dados_saida)

                # Salvar as alterações na planilha
                workbook.save(caminh_plan_saida)


                messagebox.showinfo("Sucesso", "Item completamente removido do banco de dados.")
            else:
                messagebox.showerror("Erro", "Quantidade a ser deletada é maior que a quantidade atual.")
        else:
            messagebox.showerror("Erro", "Produto não encontrado no banco de dados.")

    except Exception as e:
        messagebox.showerror("Erro", f"{str(e)}")
    finally:
        if conn:
            conn.close()

def voltar_menu():

    tela_estoque.withdraw()
    tela_menu.deiconify()

def menu_graficos():
    global tela_menu_graf
    tela_menu.withdraw()

    tela_menu_graf = CTkToplevel()
    tela_menu_graf.geometry("{}x{}+0+0".format(largura_tela, altura_tela))
    tela_menu_graf.title('Estoque - Menu - Gráficos')
    tela_menu_graf.protocol("WM_DELETE_WINDOW", sair_do_app)

    CTkLabel(master=tela_menu_graf, text="", image=img_lat).pack(expand=True, side="left")

    frame_cabecalho_graf = CTkFrame(master=tela_menu_graf, width=int(0.6 * largura_tela), height=int(0.14 * altura_tela))
    frame_cabecalho_graf.pack(expand=True, side="top")

    CTkLabel(master=frame_cabecalho_graf, text="------------------------------------- Estoque -------------------------------------", text_color="#00009C", anchor="w", justify="center", font=("Arial Bold", 24)).pack(anchor="w", pady=(0,0), padx=(0, 0))

    scrollable_frame_graf = CTkScrollableFrame(master=tela_menu_graf, width=int(0.6 * largura_tela), height=int(0.85 * altura_tela), fg_color="#ffffff")
    scrollable_frame_graf.pack(expand=True, side="right")

    frame_tela_graf = CTkFrame(master=scrollable_frame_graf, fg_color="#ffffff")
    frame_tela_graf.pack(expand=True, anchor="w", pady=(10, 0))

    CTkLabel(master=frame_tela_graf, text="Gráficos", text_color="#00009C", anchor="w", justify="center", font=("Arial Bold", 28), compound="left").pack(anchor="w", pady=(10, 0), padx=(20, 0))

    CTkLabel(master=frame_tela_graf, text="  Acompanhamento Estoque Físico:", text_color="#00009C", anchor="w", justify="center", font=("Arial Bold", 14), compound="left").pack(anchor="w", pady=(35, 0), padx=(int(0.2*frame_tela_graf.winfo_screenwidth()), 0))
    CTkButton(master=frame_tela_graf, text="Ver Gráficos", fg_color="#00009C", hover_color="#E44982", font=("Arial Bold", 12), text_color="#ffffff", width=225, command=graf_estoque).pack(anchor="w", pady=(15, 0), padx=(int(0.2*frame_tela_graf.winfo_screenwidth()), 0))

    CTkLabel(master=frame_tela_graf, text="  Voltar", text_color="#00009C", anchor="w", justify="center", font=("Arial Bold", 14), compound="left").pack(anchor="w", pady=(35, 0), padx=(int(0.24*frame_tela_graf.winfo_screenwidth()), 0))
    CTkButton(master=frame_tela_graf, text="Voltar ao Menu", fg_color="#00009C", hover_color="#E44982", font=("Arial Bold", 12), text_color="#ffffff", width=225, command=voltar_menu_visib).pack(anchor="w", pady=(15, 0), padx=(int(0.2*frame_tela_graf.winfo_screenwidth()), 0))

def graf_estoque():
    global combo_produto_stk, frame_grafico_qtd_prod,  df_estoque_visib, frame_scroll_graf_estoque
    tela_menu_graf.withdraw()

    # Conectar ao banco de dados SQLite
    conn = sqlite3.connect(caminh_banco)
    df_estoque_visib= pd.read_sql_query("SELECT code, qtd, preco, codigo_barras FROM fisico",conn)
    produtos = [item for item in df_estoque_visib['code']]

    conn.close()

    df_estoque_visib['qtd'] = df_estoque_visib['qtd'].astype(int)
    df_estoque_visib['preco'] = df_estoque_visib['preco'].astype(float)

    global tela_graf_estoque

    tela_graf_estoque = CTkToplevel()
    tela_graf_estoque.geometry("{}x{}+0+0".format(largura_tela, altura_tela))
    tela_graf_estoque.title("Estoque - Estoque Físico - Gráficos")
    tela_graf_estoque.protocol("WM_DELETE_WINDOW", sair_do_app)

    frame_scroll_graf_estoque = CTkScrollableFrame(master=tela_graf_estoque, width= int(largura_tela), height=int(altura_tela), fg_color="#ffffff")
    frame_scroll_graf_estoque.pack_propagate(0)
    frame_scroll_graf_estoque.pack(expand=True, side="right")

    frame_cabecalho_graf_estoque = ttk.Frame(frame_scroll_graf_estoque)
    frame_cabecalho_graf_estoque.grid(row=0, column=0, padx=60, pady=10)

    frame_filtro_estoque = ttk.Frame(frame_cabecalho_graf_estoque)
    frame_filtro_estoque.grid(row=0, column=0, padx=10, pady=10)

    CTkButton(master=frame_filtro_estoque, text="Voltar", fg_color="#00009C", hover_color="#E44982", font=("Arial Bold", 12), text_color="#ffffff", width=50, command=voltar_graf_estoque).grid(row=1, column=1, padx=10, pady=10)
    CTkLabel(master=frame_filtro_estoque, text="Estoque - Gráficos Estoque Físico", text_color="#00009C", anchor="w", justify="left", font=("Arial Bold", 24)).grid(row=1, column=0, padx=10, pady=10)

    produtos_unicos_stk = list(set(produtos))
    combo_produto_stk = ttk.Combobox(master=frame_filtro_estoque, values=produtos_unicos_stk)
    combo_produto_stk.grid(row=2, column=0, padx=10, pady=10)
    combo_produto_stk.set("Selecione um Produto")

    btn_filtro = CTkButton(master=frame_filtro_estoque, width=10, text="Filtrar", command=filtro_estoque_fisico)
    btn_filtro.grid(row=3, column=0, padx=10, pady=10)

    btn_limpar_filtro = CTkButton(master=frame_filtro_estoque, width=10, text="Limpar Filtro", command=limpar_filtro_estoque)
    btn_limpar_filtro.grid(row=3, column=1, padx=10, pady=10)


    # Agrupar os dados por código de produto e somar as quantidades
    df_agrupado_prod_qtd = df_estoque_visib.groupby('code')['qtd'].sum().reset_index()

    # Configurar o tamanho do gráfico
    plt.figure(figsize=(16, 6))

    # Definir uma lista de cores com base nas condições
    colors = []
    for qtd in df_agrupado_prod_qtd['qtd']:
        if qtd > 100:
            colors.append('#90EE90')  # Verde claro
        elif qtd > 60 and qtd <= 100:
            colors.append('#90EE90')  # Verde claro
        elif qtd > 30 and qtd <= 60:
            colors.append('yellow')  # Amarelo
        elif qtd <= 30:
            colors.append('#FF7F7F')  # Vermelho claro
        else:
            colors.append('blue')  # Cor padrão (azul) para valores > 100

    # Plotar as barras com as cores definidas
    bars1 = plt.bar(df_agrupado_prod_qtd['code'], df_agrupado_prod_qtd['qtd'], color=colors)

    # Adicionar rótulos e título
    plt.xlabel('Produtos')
    plt.ylabel('Quantidade')
    plt.title('Gestão Visual')
    plt.xticks(rotation=20, ha='right', fontsize=6)

    # Adicionar os valores nas barras
    for bar in bars1:
        yval = bar.get_height()
        plt.text(bar.get_x() + bar.get_width()/2, yval + 0.02 * plt.ylim()[1], 
                round(yval, 2), ha='center', va='top', fontweight='bold', color='black', fontsize=7)


    # Adicionar gráfico de barras ao frame
    frame_grafico_qtd_prod = ttk.Frame(frame_scroll_graf_estoque)
    frame_grafico_qtd_prod.grid(row=2, column=0, padx=10, pady=10, sticky="ew", columnspan=2)

    canvas_qtd_prod = FigureCanvasTkAgg(plt.gcf(), master=frame_grafico_qtd_prod)
    canvas_qtd_prod.draw()
    canvas_qtd_prod.get_tk_widget().pack(side=tkinter.TOP, fill=tkinter.BOTH, expand=1)

    plt.close()


def filtro_estoque_fisico():
    global frame_grafico_qtd_prod, df_estoque_visib
    if (combo_produto_stk.get() != "Selecione um Produto" or combo_produto_stk.get() != ""):
        frame_grafico_qtd_prod.destroy()
        produto_selecionado = str(combo_produto_stk.get())
        
        df_estoque_visib['qtd'] = df_estoque_visib['qtd'].astype(int)
        df_estoque_visib['preco'] = df_estoque_visib['preco'].astype(float)

        df_estoque_visib = df_estoque_visib[df_estoque_visib['code'] == produto_selecionado]

        # Agrupar os dados por código de produto e somar as quantidades
        df_agrupado_prod_qtd = df_estoque_visib.groupby('code')['qtd'].sum().reset_index()

        # Configurar o tamanho do gráfico
        plt.figure(figsize=(16, 6))

        # Definir uma lista de cores com base nas condições
        colors = []
        for qtd in df_agrupado_prod_qtd['qtd']:
            if qtd > 100:
                colors.append('#90EE90')  # Verde claro
            elif qtd > 60 and qtd <= 100:
                colors.append('#90EE90')  # Verde claro
            elif qtd > 30 and qtd <= 60:
                colors.append('yellow')  # Amarelo
            elif qtd <= 30:
                colors.append('#FF7F7F')  # Vermelho claro
            else:
                colors.append('blue')  # Cor padrão (azul) para valores > 100

        # Plotar as barras com as cores definidas
        bars1 = plt.bar(df_agrupado_prod_qtd['code'], df_agrupado_prod_qtd['qtd'], color=colors)

        # Adicionar rótulos e título
        plt.xlabel('Produtos')
        plt.ylabel('Quantidade')
        plt.title('Quantidade por Produto')
        plt.xticks(rotation=20, ha='right', fontsize=6)

        # Adicionar os valores nas barras
        for bar in bars1:
            yval = bar.get_height()
            plt.text(bar.get_x() + bar.get_width()/2, yval + 0.02 * plt.ylim()[1], 
                    round(yval, 2), ha='center', va='top', fontweight='bold', color='black', fontsize=7)



        # Adicionar gráfico de barras ao frame
        frame_grafico_qtd_prod = ttk.Frame(frame_scroll_graf_estoque)
        frame_grafico_qtd_prod.grid(row=2, column=0, padx=10, pady=10, sticky="ew", columnspan=2)

        canvas_qtd_prod = FigureCanvasTkAgg(plt.gcf(), master=frame_grafico_qtd_prod)
        canvas_qtd_prod.draw()
        canvas_qtd_prod.get_tk_widget().pack(side=tkinter.TOP, fill=tkinter.BOTH, expand=1)

        plt.close()


def limpar_filtro_estoque():
    global frame_grafico_qtd_prod,  df_estoque_visib
    frame_grafico_qtd_prod.destroy()
    conn = sqlite3.connect(caminh_banco)
    df_estoque_visib= pd.read_sql_query("SELECT code, qtd, preco, codigo_barras FROM fisico",conn)

    conn.close()

    df_estoque_visib['qtd'] = df_estoque_visib['qtd'].astype(int)
    df_estoque_visib['preco'] = df_estoque_visib['preco'].astype(float)

    # Agrupar os dados por código de produto e somar as quantidades
    df_agrupado_prod_qtd = df_estoque_visib.groupby('code')['qtd'].sum().reset_index()

    # Configurar o tamanho do gráfico
    plt.figure(figsize=(16, 6))

    # Definir uma lista de cores com base nas condições
    colors = []
    for qtd in df_agrupado_prod_qtd['qtd']:
        if qtd > 100:
            colors.append('#90EE90')  # Verde claro
        elif qtd > 60 and qtd <= 100:
            colors.append('#90EE90')  # Verde claro
        elif qtd > 30 and qtd <= 60:
            colors.append('yellow')  # Amarelo
        elif qtd <= 30:
            colors.append('#FF7F7F')  # Vermelho claro
        else:
            colors.append('blue')  # Cor padrão (azul) para valores > 100

    # Plotar as barras com as cores definidas
    bars1 = plt.bar(df_agrupado_prod_qtd['code'], df_agrupado_prod_qtd['qtd'], color=colors)

    # Adicionar rótulos e título
    plt.xlabel('Produtos')
    plt.ylabel('Quantidade')
    plt.title('Quantidade por Produto')
    plt.xticks(rotation=20, ha='right', fontsize=6)

    # Adicionar os valores nas barras
    for bar in bars1:
        yval = bar.get_height()
        plt.text(bar.get_x() + bar.get_width()/2, yval + 0.02 * plt.ylim()[1], 
                round(yval, 2), ha='center', va='top', fontweight='bold', color='black', fontsize=7)



    # Adicionar gráfico de barras ao frame
    frame_grafico_qtd_prod = ttk.Frame(frame_scroll_graf_estoque)
    frame_grafico_qtd_prod.grid(row=2, column=0, padx=10, pady=10, sticky="ew", columnspan=2)

    canvas_qtd_prod = FigureCanvasTkAgg(plt.gcf(), master=frame_grafico_qtd_prod)
    canvas_qtd_prod.draw()
    canvas_qtd_prod.get_tk_widget().pack(side=tkinter.TOP, fill=tkinter.BOTH, expand=1)

    plt.close()


def voltar_graf_estoque():
    tela_graf_estoque.withdraw()
    tela_menu_graf.deiconify()

def menu_acomp():
    global tela_menu_acomp_saida
    tela_menu.withdraw()

    tela_menu_acomp_saida = CTkToplevel()
    tela_menu_acomp_saida.geometry("{}x{}+0+0".format(largura_tela, altura_tela))
    tela_menu_acomp_saida.title('Estoque - Menu - Acompanhamento')
    tela_menu_acomp_saida.protocol("WM_DELETE_WINDOW", sair_do_app)

    CTkLabel(master=tela_menu_acomp_saida, text="", image=img_lat).pack(expand=True, side="left")

    frame_cabecalho_acomp = CTkFrame(master=tela_menu_acomp_saida, width=int(0.6 * largura_tela), height=int(0.14 * altura_tela))
    frame_cabecalho_acomp.pack(expand=True, side="top")

    CTkLabel(master=frame_cabecalho_acomp, text="------------------------------------- Estoque -------------------------------------", text_color="#00009C", anchor="w", justify="center", font=("Arial Bold", 24)).pack(anchor="w", pady=(0,0), padx=(0, 0))

    scrollable_frame_acomp = CTkScrollableFrame(master=tela_menu_acomp_saida, width=int(0.6 * largura_tela), height=int(0.85 * altura_tela), fg_color="#ffffff")
    scrollable_frame_acomp.pack(expand=True, side="right")

    frame_tela_acomp = CTkFrame(master=scrollable_frame_acomp, fg_color="#ffffff")
    frame_tela_acomp.pack(expand=True, anchor="w", pady=(10, 0))

    CTkLabel(master=frame_tela_acomp, text="Acompanhamento", text_color="#00009C", anchor="w", justify="center", font=("Arial Bold", 28), compound="left").pack(anchor="w", pady=(10, 0), padx=(20, 0))

    CTkLabel(master=frame_tela_acomp, text="  Histórico de Entrada:", text_color="#00009C", anchor="w", justify="center", font=("Arial Bold", 14), compound="left").pack(anchor="w", pady=(35, 0), padx=(int(0.2*frame_tela_acomp.winfo_screenwidth()), 0))
    CTkButton(master=frame_tela_acomp, text="Ver Histórico de Entrada", fg_color="#00009C", hover_color="#E44982", font=("Arial Bold", 12), text_color="#ffffff", width=225, command=acomp_entrada).pack(anchor="w", pady=(15, 0), padx=(int(0.2*frame_tela_acomp.winfo_screenwidth()), 0))

    CTkLabel(master=frame_tela_acomp, text="  Histórico de Saída:", text_color="#00009C", anchor="w", justify="center", font=("Arial Bold", 14), compound="left").pack(anchor="w", pady=(35, 0), padx=(int(0.2*frame_tela_acomp.winfo_screenwidth()), 0))
    CTkButton(master=frame_tela_acomp, text="Ver Histórico de Saída", fg_color="#00009C", hover_color="#E44982", font=("Arial Bold", 12), text_color="#ffffff", width=225, command=acomp_saida).pack(anchor="w", pady=(15, 0), padx=(int(0.2*frame_tela_acomp.winfo_screenwidth()), 0))

    CTkLabel(master=frame_tela_acomp, text="  Voltar", text_color="#00009C", anchor="w", justify="center", font=("Arial Bold", 14), compound="left").pack(anchor="w", pady=(35, 0), padx=(int(0.24*frame_tela_acomp.winfo_screenwidth()), 0))
    CTkButton(master=frame_tela_acomp, text="Voltar ao Menu", fg_color="#00009C", hover_color="#E44982", font=("Arial Bold", 12), text_color="#ffffff", width=225, command=voltar_menu_acomp).pack(anchor="w", pady=(15, 0), padx=(int(0.2*frame_tela_acomp.winfo_screenwidth()), 0))

def acomp_entrada():
    tela_menu_acomp_saida.withdraw()

    global tela_acomp_entrada, combo_mes_entrada, treeview_ent

    conn = sqlite3.connect(caminh_banco)

    df_entrada= pd.read_sql_query("SELECT * FROM entrada",conn)
    mes = [item for item in df_entrada['mes_entrada']]

    conn.close()

    tela_acomp_entrada = CTkToplevel()
    tela_acomp_entrada.geometry("{}x{}+0+0".format(largura_tela, altura_tela))
    tela_acomp_entrada.title("Estoque - Histórico Entrada")
    tela_acomp_entrada.protocol("WM_DELETE_WINDOW", sair_do_app)

    frame_cabecalho_entrada = CTkFrame(master=tela_acomp_entrada, width=int(0.6 * largura_tela), height=int(0.14 * altura_tela), fg_color="#ffffff")
    frame_cabecalho_entrada.pack(expand=True, side="top")


    frame_filtro_entrada = ttk.Frame(frame_cabecalho_entrada)
    frame_filtro_entrada.grid(row=0, column=0, padx=10, pady=10)

    meses = list(set(mes))
    combo_mes_entrada = ttk.Combobox(master=frame_filtro_entrada, values=meses)
    combo_mes_entrada.grid(row=2, column=1, padx=20, pady=10)
    combo_mes_entrada.set("Selecione um Mês")

    botao_pes = ttk.Button(frame_filtro_entrada, text="Filtrar",command=filtrar_mes_entrada)
    botao_pes.grid(row=3, column=1, padx=20, pady=10)
        
    botao_vw = ttk.Button(frame_filtro_entrada, text="Limpar Filtro",command=limpar_filtro_entrada)
    botao_vw.grid(row=2, column=2, padx=20, pady=10)

    CTkButton(master=frame_filtro_entrada, text="Voltar", fg_color="#00009C", hover_color="#E44982", font=("Arial Bold", 12), text_color="#ffffff", width=50, command=voltar_acomp_entrada).grid(row=1, column=2, padx=10, pady=10)
    CTkLabel(master=frame_filtro_entrada, text="Estoque - Histórico Entrada", text_color="#00009C", anchor="w", justify="left", font=("Arial Bold", 24)).grid(row=1, column=0, padx=10, pady=10)
    
    frame_entrada = tkinter.Frame(tela_acomp_entrada, width= int(largura_tela), height=int(altura_tela))
    frame_entrada.pack_propagate(0)
    frame_entrada.pack(expand=True, side="right")

    frame_ent_treeview = tkinter.Frame(frame_entrada, width= int(largura_tela), height=int(altura_tela))
    frame_ent_treeview.pack_propagate(0)
    frame_ent_treeview.pack(expand=True, side="right")

    # Criar a Treeview
    treeview_ent = ttk.Treeview(frame_ent_treeview, columns=("Nome", "Descrição"), show="headings")
    treeview_ent.heading("Nome", text="Nome")
    treeview_ent.heading("Descrição", text="Descrição")

    # Adicionar Scrollbars
    scrollbar_y = ttk.Scrollbar(frame_ent_treeview, orient="vertical", command=treeview_ent.yview)
    scrollbar_y.pack(side="right", fill="y")
    treeview_ent.configure(yscrollcommand=scrollbar_y.set)

    scrollbar_x = ttk.Scrollbar(frame_ent_treeview, orient="horizontal", command=treeview_ent.xview)
    scrollbar_x.pack(side="bottom", fill="x")
    treeview_ent.configure(xscrollcommand=scrollbar_x.set)

    try:
            conn = sqlite3.connect(caminh_banco)
            cursor = conn.cursor()

            cursor.execute("SELECT code, descricao, qtd, un, preco, codigo_barras, data_entrada FROM entrada")
            rows = cursor.fetchall()

            treeview_ent.delete(*treeview_ent.get_children())

            cols = ["CODE", "DESCRIÇÃO", "QUANTIDADE", "UNIDADE","PREÇO","CÓDIGO DE BARRAS","DATA ENTRADA"]
            treeview_ent["columns"] = cols

            for col_name in cols:
                treeview_ent.heading(col_name, text=col_name)
                treeview_ent.column(col_name, anchor=tkinter.CENTER)

            for row in rows:
                    treeview_ent.insert('', tkinter.END, values=row)

            treeview_ent.pack(expand=True, fill="both")  
            messagebox.showinfo("Aviso!","Dados carregados com sucesso!")
    except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar dados do banco de dados: {str(e)}")
    finally:
            if conn:
                conn.close()

def filtrar_mes_entrada():
    mes_pesq = combo_mes_entrada.get()

    try:
        conn = sqlite3.connect(caminh_banco)
        cursor = conn.cursor()

        cursor.execute(f"SELECT code, descricao, qtd, un, preco, codigo_barras, data_entrada FROM entrada WHERE mes_entrada = '{mes_pesq}'")
        rows = cursor.fetchall()

        treeview_ent.delete(*treeview_ent.get_children())

        cols = ["CODE", "DESCRIÇÃO", "QUANTIDADE", "UNIDADE","PREÇO","CÓDIGO DE BARRAS","DATA ENTRADA"]
        treeview_ent["columns"] = cols

        for col_name in cols:
            treeview_ent.heading(col_name, text=col_name)
            treeview_ent.column(col_name, anchor=tkinter.CENTER)


        for row in rows:
                treeview_ent.insert('', tkinter.END, values=row)

    except Exception as e:
        messagebox.showerror("Erro", f"Item não encontrado: {str(e)}")
    finally:
        if conn:
            conn.close()

def limpar_filtro_entrada():
    try:
            conn = sqlite3.connect(caminh_banco)
            cursor = conn.cursor()

            cursor.execute("SELECT code, descricao, qtd, un, preco, codigo_barras, data_entrada FROM entrada")
            rows = cursor.fetchall()

            treeview_ent.delete(*treeview_ent.get_children())

            cols = ["CODE", "DESCRIÇÃO", "QUANTIDADE", "UNIDADE","PREÇO","CÓDIGO DE BARRAS","DATA ENTRADA"]
            treeview_ent["columns"] = cols

            for col_name in cols:
                treeview_ent.heading(col_name, text=col_name)
                treeview_ent.column(col_name, anchor=tkinter.CENTER)

            for row in rows:
                    treeview_ent.insert('', tkinter.END, values=row)

            treeview_ent.pack(expand=True, fill="both")  
            messagebox.showinfo("Aviso!","Dados carregados com sucesso!")
    except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar dados do banco de dados: {str(e)}")
    finally:
            if conn:
                conn.close()

def acomp_saida():
    tela_menu_acomp_saida.withdraw()

    global tela_acomp_saida, combo_mes_saida, treeview_sai

    conn = sqlite3.connect(caminh_banco)

    df_saida= pd.read_sql_query("SELECT * FROM saida",conn)
    mes = [item for item in df_saida['mes_saida']]


    conn.close()

    tela_acomp_saida = CTkToplevel()
    tela_acomp_saida.geometry("{}x{}+0+0".format(largura_tela, altura_tela))
    tela_acomp_saida.title("Estoque - Histórico Saída")
    tela_acomp_saida.protocol("WM_DELETE_WINDOW", sair_do_app)

    frame_cabecalho_saida = CTkFrame(master=tela_acomp_saida, width=int(0.6 * largura_tela), height=int(0.14 * altura_tela), fg_color="#ffffff")
    frame_cabecalho_saida.pack(expand=True, side="top")

    frame_filtro_saida = ttk.Frame(frame_cabecalho_saida)
    frame_filtro_saida.grid(row=0, column=0, padx=10, pady=10)

    meses = list(set(mes))
    combo_mes_saida = ttk.Combobox(master=frame_filtro_saida, values=meses)
    combo_mes_saida.grid(row=2, column=1, padx=20, pady=10)
    combo_mes_saida.set("Selecione um Mês")

    botao_pes = ttk.Button(frame_filtro_saida, text="Filtrar",command=filtrar_mes_saida)
    botao_pes.grid(row=3, column=1, padx=20, pady=10)
        
    botao_vw = ttk.Button(frame_filtro_saida, text="Limpar Filtro",command=limpar_filtro_saida)
    botao_vw.grid(row=2, column=2, padx=20, pady=10)


    CTkButton(master=frame_filtro_saida, text="Voltar", fg_color="#00009C", hover_color="#E44982", font=("Arial Bold", 12), text_color="#ffffff", width=50, command=voltar_acomp_saida).grid(row=1, column=2, padx=10, pady=10)
    CTkLabel(master=frame_filtro_saida, text="Estoque - Histórico Saída", text_color="#00009C", anchor="w", justify="left", font=("Arial Bold", 24)).grid(row=1, column=0, padx=10, pady=10)
    
    frame_saida = tkinter.Frame(tela_acomp_saida, width= int(largura_tela), height=int(altura_tela))
    frame_saida.pack_propagate(0)
    frame_saida.pack(expand=True, side="right")

    frame_sai_treeview = tkinter.Frame(frame_saida, width= int(largura_tela), height=int(altura_tela))
    frame_sai_treeview.pack_propagate(0)
    frame_sai_treeview.pack(expand=True, side="right")

    # Criar a Treeview
    treeview_sai = ttk.Treeview(frame_sai_treeview, columns=("Nome", "Descrição"), show="headings")
    treeview_sai.heading("Nome", text="Nome")
    treeview_sai.heading("Descrição", text="Descrição")

    # Adicionar Scrollbars
    scrollbar_y = ttk.Scrollbar(frame_sai_treeview, orient="vertical", command=treeview_sai.yview)
    scrollbar_y.pack(side="right", fill="y")
    treeview_sai.configure(yscrollcommand=scrollbar_y.set)

    scrollbar_x = ttk.Scrollbar(frame_sai_treeview, orient="horizontal", command=treeview_sai.xview)
    scrollbar_x.pack(side="bottom", fill="x")
    treeview_sai.configure(xscrollcommand=scrollbar_x.set)

    try:
            conn = sqlite3.connect(caminh_banco)
            cursor = conn.cursor()

            cursor.execute("SELECT code, descricao, qtd, un, preco, codigo_barras, data_saida FROM saida")
            rows = cursor.fetchall()

            treeview_sai.delete(*treeview_sai.get_children())

            cols = ["CODE", "DESCRIÇÃO", "QUANTIDADE", "UNIDADE","PREÇO","CÓDIGO DE BARRAS","DATA SAÍDA"]
            treeview_sai["columns"] = cols

            for col_name in cols:
                treeview_sai.heading(col_name, text=col_name)
                treeview_sai.column(col_name, anchor=tkinter.CENTER)

            for row in rows:
                    treeview_sai.insert('', tkinter.END, values=row)

            treeview_sai.pack(expand=True, fill="both")  
            messagebox.showinfo("Aviso!","Dados carregados com sucesso!")
    except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar dados do banco de dados: {str(e)}")
    finally:
            if conn:
                conn.close()

def filtrar_mes_saida():
    mes_pesq = combo_mes_saida.get()

    try:
        conn = sqlite3.connect(caminh_banco)
        cursor = conn.cursor()

        cursor.execute(f"SELECT code, descricao, qtd, un, preco, codigo_barras, data_saida FROM saida WHERE mes_saida = '{mes_pesq}'")
        rows = cursor.fetchall()

        treeview_sai.delete(*treeview_sai.get_children())

        cols = ["CODE", "DESCRIÇÃO", "QUANTIDADE", "UNIDADE","PREÇO","CÓDIGO DE BARRAS","DATA SAIDA"]
        treeview_sai["columns"] = cols

        for col_name in cols:
            treeview_sai.heading(col_name, text=col_name)
            treeview_sai.column(col_name, anchor=tkinter.CENTER)


        for row in rows:
                treeview_sai.insert('', tkinter.END, values=row)

    except Exception as e:
        messagebox.showerror("Erro", f"Item não encontrado: {str(e)}")
    finally:
        if conn:
            conn.close()

def limpar_filtro_saida():
    try:
            conn = sqlite3.connect(caminh_banco)
            cursor = conn.cursor()

            cursor.execute("SELECT code, descricao, qtd, un, preco, codigo_barras, data_saida FROM saida")
            rows = cursor.fetchall()

            treeview_sai.delete(*treeview_sai.get_children())

            cols = ["CODE", "DESCRIÇÃO", "QUANTIDADE", "UNIDADE","PREÇO","CÓDIGO DE BARRAS","DATA SAIDA"]
            treeview_sai["columns"] = cols

            for col_name in cols:
                treeview_sai.heading(col_name, text=col_name)
                treeview_sai.column(col_name, anchor=tkinter.CENTER)

            for row in rows:
                    treeview_sai.insert('', tkinter.END, values=row)

            treeview_sai.pack(expand=True, fill="both")  
            messagebox.showinfo("Aviso!","Dados carregados com sucesso!")
    except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar dados do banco de dados: {str(e)}")
    finally:
            if conn:
                conn.close()

def voltar_acomp_entrada():
    tela_acomp_entrada.withdraw()
    tela_menu_acomp_saida.deiconify()

def voltar_acomp_saida():
    tela_acomp_saida.withdraw()
    tela_menu_acomp_saida.deiconify()

def voltar_menu_visib():
    tela_menu_graf.withdraw()
    tela_menu.deiconify()

def voltar_menu_acomp():
    tela_menu_acomp_saida.withdraw()
    tela_menu.deiconify()

def voltar_encontrar_item():
    tela_layout.withdraw()
    tela_menu.deiconify()

def encontrar_item():
    global tela_layout, code_search_entry, frame_layout
    tela_menu.withdraw()

    tela_layout = CTkToplevel()
    tela_layout.geometry("{}x{}+0+0".format(largura_tela, altura_tela))
    tela_layout.title('Estoque - Layout')
    tela_layout.protocol("WM_DELETE_WINDOW", sair_do_app)

    CTkLabel(master=tela_layout, text="", image=img_lat).pack(expand=True, side="left")

    frame_cabecalho_layout = CTkFrame(master=tela_layout, width=int(0.6 * largura_tela), height=int(0.14 * altura_tela))
    frame_cabecalho_layout.pack(expand=True, side="top")

    CTkButton(master=frame_cabecalho_layout, text="Voltar", fg_color="#00009C", hover_color="#E44982", font=("Arial Bold", 12), text_color="#ffffff", width=50, command=voltar_encontrar_item).pack(anchor="w", pady=(0,0), padx=(int(0.55*frame_cabecalho_layout.winfo_screenwidth()), 0))
    CTkLabel(master=frame_cabecalho_layout, text="Estoque - Encontrar item", text_color="#00009C", anchor="w", justify="left", font=("Arial Bold", 18)).pack(anchor="w", pady=(10,0), padx=(0, 0))

    frame_lay_scroll = CTkScrollableFrame(master=tela_layout, width= int(0.6 * largura_tela), height=int(0.85 * altura_tela), fg_color="#ffffff")
    frame_lay_scroll.pack(expand=True, side="right")

    frame_layout = ttk.Frame(frame_lay_scroll)
    frame_layout.grid(row=0, column=0, padx=10, pady=20)
    frame_layout.pack(expand=True)

    CTkLabel(master=frame_layout, text="----------------Estoque - Encontrar item----------------", text_color="#ffffff", anchor="w", justify="left", font=("Arial Bold", 18)).grid(row=1, column=1, padx=10, pady=20)
    code_search_entry = CTkEntry(master=frame_layout, width=225, fg_color="#EEEEEE", border_color="#00009C", border_width=1, text_color="#000000")
    code_search_entry.grid(row=2, column=1, padx=10, pady=20)
    CTkButton(master=frame_layout, text="Procurar", fg_color="#00009C", hover_color="#E44982", font=("Arial Bold", 12), text_color="#ffffff", width=50, command=ler_layout).grid(row=3, column=1, padx=10, pady=20)

def ler_layout():
    codigo_para_encontrar = code_search_entry.get()
    # Conectar ao banco de dados SQLite
    conexao = sqlite3.connect(caminh_banco)

    # Criar um cursor para executar comandos SQL
    cursor = conexao.cursor()

    cursor.execute("SELECT descricao FROM fisico WHERE code = ?", (codigo_para_encontrar,))
    descricao = cursor.fetchone()[0]
    conexao.close()

    df_layout = pd.read_excel(caminh_layout, sheet_name="LAYOUT")
    df_layout = df_layout[df_layout['DESCRIÇÃO'] == descricao]

    lay = df_layout['LAYOUT'].iloc[0]
    local = df_layout['LOCAL'].iloc[0]


    CTkLabel(master=frame_layout, text="Descrição:", text_color="#00009C", anchor="w", justify="left", font=("Arial Bold", 18)).grid(row=4, column=1, padx=10, pady=20)
    CTkLabel(master=frame_layout, text=f"{descricao}", text_color="#00009C", anchor="w", justify="left", font=("Arial Bold", 18)).grid(row=5, column=1, padx=10, pady=20)

    CTkLabel(master=frame_layout, text="Onde Está:", text_color="#00009C", anchor="w", justify="left", font=("Arial Bold", 18)).grid(row=6, column=1, padx=10, pady=20)
    CTkLabel(master=frame_layout, text=f"{lay}", text_color="#00009C", anchor="w", justify="left", font=("Arial Bold", 18)).grid(row=7, column=1, padx=10, pady=20)

    CTkLabel(master=frame_layout, text="Local:", text_color="#00009C", anchor="w", justify="left", font=("Arial Bold", 18)).grid(row=8, column=1, padx=10, pady=20)
    CTkLabel(master=frame_layout, text=f"{local}", text_color="#00009C", anchor="w", justify="left", font=("Arial Bold", 18)).grid(row=9, column=1, padx=10, pady=20)


def mostrar_senha():
    stat = check_senha_var.get()
    if stat == "on":
        passw.configure(show='')
    else:
        passw.configure(show='*')

check_senha_var = tkinter.StringVar(master=tela_login)

CTkLabel(master=tela_login, text="", image=img_lat).pack(expand=True, side="left")

frame_login = CTkFrame(master=tela_login, width= int(0.6 * largura_tela), height=int(altura_tela), fg_color="#ffffff")
frame_login.pack_propagate(0)
frame_login.pack(expand=True, side="right")

CTkLabel(master=frame_login, text="Bem Vindo(a)!", text_color="#00009C", anchor="w", justify="left", font=("Arial Bold", 24)).pack(anchor="w", pady=(20, 5), padx=(25, 0))
CTkLabel(master=frame_login, text="Realize o Login", text_color="#7E7E7E", anchor="w", justify="left", font=("Arial Bold", 12)).pack(anchor="w", padx=(25, 0))

CTkLabel(master=frame_login, text=" Login:", text_color="#00009C", anchor="w", justify="left", font=("Arial Bold", 14), image=icone_user, compound="left").pack(anchor="w", pady=(100, 0), padx=(int(0.2*frame_login.winfo_screenwidth()), 0))
logn = CTkEntry(master=frame_login, width=225, fg_color="#EEEEEE", border_color="#00009C", border_width=1, text_color="#000000")
logn.pack(anchor="w", padx=(int(0.2*frame_login.winfo_screenwidth()), 0))

CTkLabel(master=frame_login, text=" Senha:", text_color="#00009C", anchor="w", justify="left", font=("Arial Bold", 14), image=icone_senha, compound="left").pack(anchor="w", pady=(21, 0), padx=(int(0.2*frame_login.winfo_screenwidth()), 0))
passw = CTkEntry(master=frame_login, width=225, fg_color="#EEEEEE", border_color="#00009C", border_width=1, text_color="#000000", show="*")
passw.pack(anchor="w", padx=(int(0.2*frame_login.winfo_screenwidth()), 0))

checkbox = CTkCheckBox(master=frame_login, text=" Mostrar Senha:", text_color="#00009C", font=("Arial Bold", 12),command=mostrar_senha,variable=check_senha_var, onvalue="on", offvalue="off").pack(anchor="w", pady=(25, 0), padx=(int(0.2*frame_login.winfo_screenwidth()), 0))

CTkButton(master=frame_login, text="Login", fg_color="#00009C", hover_color="#E44982", font=("Arial Bold", 12), text_color="#ffffff", width=225, command=autenticacao).pack(anchor="w", pady=(40, 0), padx=(int(0.2*frame_login.winfo_screenwidth()), 0))
tela_login.protocol("WM_DELETE_WINDOW", sair_do_app)
tela_login.mainloop()
